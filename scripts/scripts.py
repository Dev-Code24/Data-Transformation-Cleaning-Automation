import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
import math
import numpy as np

def extract_data(file_path, file_type='excel'):
    """Extracts data from an Excel (.xlsx) file."""
    if file_type == 'excel':
        return pd.read_excel(file_path)
    else:
        raise ValueError("Unsupported file type")


def load_data(df, output_path):
    """Saves the transformed data to an Excel (.xlsx) file."""
    df.to_excel(output_path, index=False)

def transform_data(df):
    df.columns = [col.lower().replace(" ", "_") for col in df.columns]
    df.drop_duplicates(inplace=True)
    
    df['age'].fillna(math.ceil(df['age'].mean()), inplace=True)
    df['score'].fillna(df['score'].mean(), inplace=True)
    return df

def generate_report(df, output_path, threshold_value=50, chart_column="category"):
    """Generates an Excel report with conditional formatting and a pie chart for a specified categorical column."""
    
    # Step 1: Save the DataFrame to an Excel file
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Report', index=False)
    
    # Step 2: Load the workbook and sheet for additional modifications
    wb = load_workbook(output_path)
    ws_main = wb['Report']
    
    # Step 3: Apply conditional formatting for values below threshold
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    columns_to_check = ["age", "score", "purchase_amount"]  # Customize as needed
    col_indexes = [cell.column for cell in ws_main[1] if cell.value in columns_to_check]
    
    for row in ws_main.iter_rows(min_row=2, max_row=ws_main.max_row):
        for col_index in col_indexes:
            cell = row[col_index - 1]
            if isinstance(cell.value, (int, float)) and cell.value < threshold_value:
                cell.fill = red_fill

    # Step 4: Add a pie chart for the specified column if it's categorical
    if chart_column in df.columns and df[chart_column].dtype == 'object':
        # Calculate value counts for the categorical column
        value_counts = df[chart_column].value_counts()
        
        # Add these counts to a new sheet for chart data
        ws_chart_data = wb.create_sheet(title=f"{chart_column.capitalize()} Distribution Data")
        ws_chart_data.append([chart_column, "Count"])  # Header
        for category, count in value_counts.items():
            ws_chart_data.append([category, count])
        
        # Define references for the pie chart
        labels = Reference(ws_chart_data, min_col=1, min_row=2, max_row=1 + len(value_counts))
        data = Reference(ws_chart_data, min_col=2, min_row=1, max_row=1 + len(value_counts))

        # Initialize and configure the pie chart
        pie = PieChart()
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = f"{chart_column.capitalize()} Distribution"
        
        # Set data labels to show category names and percentages
        pie.dLbls = DataLabelList()  # Initialize the data label list
        pie.dLbls.showCatName = True  # Show category names on the chart
        pie.dLbls.showVal = True      # Show values on the chart
        pie.dLbls.showPercent = True  # Show percentages

        # Add the pie chart to the main report sheet
        ws_main.add_chart(pie, "M10")  # Position the chart in cell M10 of the Report sheet
    
    # Step 5: Save workbook with all changes
    wb.save(output_path)
    print(f"Report generated with conditional formatting and '{chart_column}' column pie chart.")
